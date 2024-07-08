from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string,get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import logging
import os
############################################################################################
uliate_strength = 530  #Mega pascal
yield_stress = 490 #Mega pascal
a_width = 600 #m
b_length = 400 #m
t_thicknes = 8.5 #m
mu = 0.34 #unit less
E_B_basis = 65959.11 #65960.68 #65959.11
m_wave = 3              ###different for all 
n_wave= 1
pi = 3.14159


b_stringer = 37.5 #mm
t_stringer = 2 #mm

r_qyro = 6.9014403 #10.3609        ###difernet for all
lambda_crit = 87.812416
I_comb = 50923.234


DIM1 = 70
DIM2 = 50
DIM3 = 3
DIM4 = 3

width_web = DIM2-DIM3

filename_panels= '1_2_2D_3D'
filename_stringers= '1_2_1D'
result_file = 'Exercise_1_1_test.xlsx'
############################################################################################


os.chdir("..")
badir =  os. getcwd()  
os.chdir("results_1_2")
femdir = os.getcwd()
# entered the FEM save space


cases = ['LC1','LC2','LC3',]
LC1dir ={}
LC2dir ={}
LC3dir ={}
maindir = {}
maindir ['LC1'] = LC1dir
maindir ['LC2'] = LC2dir
maindir ['LC3'] = LC3dir

for i in range (len(maindir)):
    maindir [f'LC{i+1}'] [f'Stringer {cases[i]}'] = {}
    maindir [f'LC{i+1}'] [f'vonMieses {cases[i]}'] = {}
    maindir [f'LC{i+1}'] [f'XX {cases[i]}'] = {}
    maindir [f'LC{i+1}'] [f'YY {cases[i]}'] = {}
    maindir [f'LC{i+1}'] [f'XY {cases[i]}'] = {}

for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'Stringer'  not in  LoadTable: 
            maindir [LoadCase] [LoadTable].update ({'panel1':{},'panel2':{},'panel3':{},'panel4':{},'panel5':{}})
        if 'Stringer'  in  LoadTable: 
            maindir [LoadCase] [LoadTable].update ({'stringer1':{},'stringer2':{},'stringer3':{},'stringer4':{}})

                  


'''panel_file = 'Panels'
stringer_file = 'Stringer'

for i in range(1,len(cases)+1):
    file_path = os.path.join(femdir,panel_file)
    wb = load_workbook(file_path)
    columns = 6
    ws = wb[f'{panel_file}']
    if i == 1:
        row =12 
        for j in range(row,row+30):
            id = {}
            id ['start_value'] = ws.cell(row=j, column=columns).value
            value = ws.cell(row=j, column=1).value
            elementid [value] = id
'''


print ('hello')

#___ getting the data V2 ___
file_path = os.path.join(femdir, f'{filename_panels}.xlsx')
wb = load_workbook(file_path)
ws = wb[filename_panels]

column = 6
row = 6

for elements in cases:
    value = 0
    for i in range (1,6):
        row += 6
        for j in range(0,6):
            value += 1
            same = str(value)
            id = {}
            id ['start_value'] =  ws.cell(row=row+j,column=column).value
            maindir [f'{elements}'] [f'XX {elements}'] [f'panel{i}'] [same]= id
            id = {}
            id ['start_value'] =  ws.cell(row=row+j,column=column+1).value
            maindir [f'{elements}'] [f'XY {elements}'] [f'panel{i}'] [same] = id
            id = {}
            id ['start_value'] =  ws.cell(row=row+j,column=column+2).value
            maindir [f'{elements}'] [f'YY {elements}'] [f'panel{i}'] [same] = id
            id = {}
            id ['start_value'] =  ws.cell(row=row+j,column=column+3).value
            maindir [f'{elements}'] [f'vonMieses {elements}'] [f'panel{i}'] [same] = id


file_path = os.path.join(femdir, f'{filename_stringers}.xlsx')
wb = load_workbook(file_path)
ws = wb[filename_stringers]
column = 5
row = 9

for elements in cases:
    value = 36
    for i in range (1,5):
        row += 3
        value += 3
        for j in range(0,3):
            value += 1
            same = str(value)
            id = {}
            id ['start_value'] =  ws.cell(row=row+j,column=column).value
            maindir [f'{elements}'] [f'Stringer {elements}'] [f'stringer{i}'] [same]= id
           

'''#___ getting the data ___
for i in range(1,len(cases)+1):
    for filename in os.listdir(femdir):
        if cases[i-1] in filename and '.xlsx' in filename:
            file_path = os.path.join(femdir, filename)
            filename = filename[:-5]
            wb = load_workbook(file_path)
            ws = wb[f'{filename}']
            elementid = {}
            panel_old = None
            for j in range (1,ws.max_row+1) :
                value = ws.cell(row=j, column=1).value
                if  type(value) != int: continue
                panel = ws.cell(row=j, column=4).value
                if panel_old != panel and panel_old is not None : 
                    maindir [f'LC{i}'] [f'{filename}'] [f'{panel_old}'] = elementid
                    elementid = {}
                id = {}
                id ['start_value'] = ws.cell(row=j, column=5).value
                value =  str(value)
                elementid [value] = id
                panel_old = panel
                
            maindir [f'LC{i}'] [f'{filename}'] [f'{panel}'] = elementid'''
                
                    
print('hekko')

#___ RF_strength ___
for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        for Pannel in maindir[LoadCase] [LoadTable]:
            for ElementID,Element in maindir[LoadCase] [LoadTable] [Pannel].items():

                if 'Stringer' in LoadTable or 'vonMieses' in LoadTable:
                    absolute_working_stress = abs(Element ['start_value'])
                    RF_strength = uliate_strength/ (1.5 * absolute_working_stress)
                    Element ['RF_strength'] = RF_strength
                    

#___ averaged_load ___
for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'XX' in LoadTable or 'YY' in LoadTable or 'XY' in LoadTable:
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                K = 0
                for ElementID,Element in maindir[LoadCase] [LoadTable] [Panel].items():
                    M = 40000 * Element ['start_value']
                    K = K + M
                result = K/(6*40000)
                specificPanels ['averaged_load'] = result

beta_LC1 ={}
beta_LC2 ={}
beta_LC3 ={}

m_LC1 = [2,2,2,2,2]
m_LC2 = [1,1,1,1,1]
m_LC3 = [2,2,2,2,2]
i = 0

k_biax_LC1 = {}
k_biax_LC2 = {}
k_biax_LC3 = {}
#___ biaxial stress ___
alpha = a_width/b_length
sigma_e = ((E_B_basis * (pi**2))/(12*(1-mu**2))) * ((t_thicknes/b_length)**2)

for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'XX' in LoadTable :
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                if i == 5 : i = 0
                sigma_x = specificPanels ['averaged_load']
                sigma_y = maindir [LoadCase][f'YY {LoadCase}'] [Panel] ['averaged_load']              
                beta = sigma_y /sigma_x
                if 'LC1' in LoadCase:
                    beta_LC1 [Panel] = beta
                if 'LC2' in LoadCase:
                    beta_LC2 [Panel] = beta
                if 'LC3' in LoadCase:
                    beta_LC3 [Panel] = beta
                specificPanels ['beta'] = beta

                if 'LC1' in LoadCase:
                    k_sigma = ((m_LC1[i]**2 + n_wave**2 * alpha**2)**2)/(alpha**2 * (m_LC1[i]**2 + beta * n_wave**2 * alpha**2))
                    sigma_crit = sigma_e * k_sigma
                    specificPanels ['sigma_crit'] = sigma_crit
                    specificPanels ['k_sigma'] = k_sigma
                    k_biax_LC1 [Panel] = k_sigma
                if 'LC2' in LoadCase:
                    k_sigma = ((m_LC2[i]**2 + n_wave**2 * alpha**2)**2)/(alpha**2 * (m_LC2[i]**2 + beta * n_wave**2 * alpha**2))
                    sigma_crit = sigma_e * k_sigma
                    specificPanels ['sigma_crit'] = sigma_crit
                    specificPanels ['k_sigma'] = k_sigma
                    k_biax_LC2 [Panel] = k_sigma
                if 'LC3' in LoadCase:
                    k_sigma = ((m_LC3[i]**2 + n_wave**2 * alpha**2)**2)/(alpha**2 * (m_LC3[i]**2 + beta * n_wave**2 * alpha**2))
                    sigma_crit = sigma_e * k_sigma
                    specificPanels ['sigma_crit'] = sigma_crit
                    specificPanels ['k_sigma'] = k_sigma
                    k_biax_LC3 [Panel] = k_sigma
                i += 1

                


#___ shear stress ___

sigma_e = ((E_B_basis * (pi**2))/(12*(1-mu**2))) * ((t_thicknes/b_length)**2)
if alpha >= 1:
    k_shear = 5.35 + (4/(alpha**2))
else:
    k_shear = 4 + (5.35/(alpha**2))
tau_crit = k_shear * sigma_e

for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if  'XY' in LoadTable:
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                specificPanels ['tau_crit'] = tau_crit


#___ RF Biaxial ___
for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'XX' in LoadTable :
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                sigma_crit = abs(specificPanels ['sigma_crit'])
                averaged_load = abs(specificPanels ['averaged_load'])
                RF_biaxial = (sigma_crit/ (1.5 * averaged_load))
                specificPanels ['RF_biaxial'] = RF_biaxial


#___ RF Shear ___
for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'XY' in LoadTable :
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                tau_crit = specificPanels ['tau_crit']
                averaged_load = specificPanels ['averaged_load']
                RF_shear = (tau_crit/ (1.5 *averaged_load))
                specificPanels ['RF_shear'] = RF_shear


#___ RF Comb ___
RF_LC1 ={}
RF_LC2 ={}
RF_LC3 ={}

for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'XX' in LoadTable :
            for Panel,specificPanels in maindir[LoadCase] [LoadTable].items():
                RF_biaxial = specificPanels ['RF_biaxial']
                RF_shear = maindir [LoadCase][f'XY {LoadCase}'] [Panel] ['RF_shear']    
                RF_comb = 1/((1/RF_biaxial)+(1/RF_shear)**2)
                specificPanels ['RF_comb'] = RF_comb
                if 'LC1' in LoadCase:
                    RF_LC1 [Panel] = RF_comb
                if 'LC2' in LoadCase:
                    RF_LC2 [Panel] = RF_comb
                if 'LC3' in LoadCase:
                    RF_LC3 [Panel] = RF_comb




#___ cripeling depression ___
K_crip = 0.41
x_crip = (b_stringer/t_stringer) * ((yield_stress/(K_crip*E_B_basis))**0.5)
alpha_crimp = 0.69/(x_crip**0.75)             
sigma_crip = alpha_crimp * yield_stress   

#___ BUCKling ___
#69 nice
if sigma_crip < yield_stress : 
    sigma_cutoff = sigma_crip 
else: sigma_cutoff = yield_stress

lambda_euler = (1 *  a_width)/r_qyro

sigma_euler_johnson = sigma_cutoff - ( (1/E_B_basis) * ((sigma_cutoff/(2*pi))**2) * (lambda_euler**2))
sigma_euler = ((pi**2) * (E_B_basis))/(lambda_euler**2)

if sigma_euler > sigma_euler_johnson : 
    sigma_buckel = sigma_euler 
else: sigma_buckel = sigma_euler_johnson

if sigma_crip < sigma_buckel : 
    sigma_buckel_crit = sigma_crip 
else: sigma_buckel_crit = sigma_buckel


pups = (DIM1 * DIM3 + (width_web * DIM4))
V_one_pbar = 200 * pups
V_one_pshell = 200*200*t_thicknes
V_one_column = (3*V_one_pbar)+(6*V_one_pshell)

help = 0

stringer_sigma_LC1 = {}
stringer_sigma_LC2 = {}
stringer_sigma_LC3 = {}

RF_stringer_LC1 = {}
RF_stringer_LC2 = {}
RF_stringer_LC3 = {}

for LoadCase in maindir:
    for LoadTable in maindir[LoadCase]:
        if 'Stringer' in LoadTable :
            for Stringer,specificStringer in maindir[LoadCase] [LoadTable].items():
                if '1' in Stringer : help = 1
                if '2' in Stringer : help = 2
                if '3' in Stringer : help = 3
                if '4' in Stringer : help = 4
                K = 0
                L = 0
                for ElementID,Element in maindir[LoadCase] [LoadTable] [Stringer].items():
                    M = Element ['start_value']
                    K = K + M
                specificStringer ['sum_of_loads_stringer'] = K
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help}'] [f'{((help-1)*6)+4}'] ['start_value']
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help}'] [f'{((help-1)*6)+5}'] ['start_value']
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help}'] [f'{((help-1)*6)+6}'] ['start_value']
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help+1}'] [f'{(help*6)+1}'] ['start_value']
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help+1}'] [f'{(help*6)+2}'] ['start_value']
                L = L + maindir[LoadCase] [f'XX {LoadCase}'] [f'panel{help+1}'] [f'{(help*6)+3}'] ['start_value']
                specificStringer ['sum_of_loads_panels'] = L
                sigma_buckel_comb = ((V_one_pbar * specificStringer ['sum_of_loads_stringer']) + (V_one_pshell * specificStringer ['sum_of_loads_panels']))/V_one_column
                RF_stringer_comb =abs(sigma_buckel_crit/(1.5 * sigma_buckel_comb))
                if 'LC1' in LoadCase:
                    stringer_sigma_LC1 [Stringer] = sigma_buckel_comb 
                if 'LC2' in LoadCase:
                    stringer_sigma_LC2 [Stringer] = sigma_buckel_comb
                if 'LC3' in LoadCase:
                    stringer_sigma_LC3 [Stringer] = sigma_buckel_comb
                specificStringer ['RF_stringer_comb'] = RF_stringer_comb
                if 'LC1' in LoadCase:
                    RF_stringer_LC1 [Stringer] = RF_stringer_comb 
                if 'LC2' in LoadCase:
                    RF_stringer_LC2 [Stringer] = RF_stringer_comb
                if 'LC3' in LoadCase:
                    RF_stringer_LC3 [Stringer] = RF_stringer_comb

#___ save ___
os.chdir("../99_Abgabe")
wb = load_workbook (result_file)
ws = wb['Tabelle1']
 

column = 2
row = 72


#___ columbe buckling___
for i in range(1,5):
    ws.cell(row=row+i,column=column).value = stringer_sigma_LC1 [f'stringer{i}']
    ws.cell(row=row+i,column=column+1).value = sigma_crip
    ws.cell(row=row+i,column=column+2).value = RF_stringer_LC1 [f'stringer{i}']

column +=5
for i in range(1,5):
    ws.cell(row=row+i,column=column).value = stringer_sigma_LC2 [f'stringer{i}']
    ws.cell(row=row+i,column=column+1).value = sigma_crip
    ws.cell(row=row+i,column=column+2).value = RF_stringer_LC2 [f'stringer{i}']

column +=5
for i in range(1,5):
    ws.cell(row=row+i,column=column).value = stringer_sigma_LC3 [f'stringer{i}']
    ws.cell(row=row+i,column=column+1).value = sigma_crip
    ws.cell(row=row+i,column=column+2).value = RF_stringer_LC3 [f'stringer{i}']


#___ i'dont know why values___
column = 2
row = 79
for i in range(1,5):
    ws.cell(row=row+i,column=column).value = I_comb
    ws.cell(row=row+i,column=column+1).value = r_qyro
    ws.cell(row=row+i,column=column+2).value = lambda_euler
    ws.cell(row=row+i,column=column+3).value = lambda_crit

#___ panel BUCKling ___
column = 2
row = 63
for i in range(1,6):
    ws.cell(row=row+i,column=column).value =  maindir ['LC1']['XX LC1'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+1).value = maindir ['LC1']['YY LC1'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+2).value = maindir ['LC1']['XY LC1'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+3).value = k_shear
    ws.cell(row=row+i,column=column+4).value = k_biax_LC1 [f'panel{i}']
    ws.cell(row=row+i,column=column+5).value = RF_LC1 [f'panel{i}']

column = 10
for i in range(1,6):
    ws.cell(row=row+i,column=column).value =  maindir ['LC2']['XX LC2'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+1).value = maindir ['LC2']['YY LC2'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+2).value = maindir ['LC2']['XY LC2'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+3).value = k_shear
    ws.cell(row=row+i,column=column+4).value = k_biax_LC2 [f'panel{i}']
    ws.cell(row=row+i,column=column+5).value = RF_LC2 [f'panel{i}']

column = 18
for i in range(1,6):
    ws.cell(row=row+i,column=column).value =  maindir ['LC3']['XX LC3'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+1).value = maindir ['LC3']['YY LC3'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+2).value = maindir ['LC3']['XY LC3'] [f'panel{i}'] ['averaged_load'] 
    ws.cell(row=row+i,column=column+3).value = k_shear
    ws.cell(row=row+i,column=column+4).value = k_biax_LC3 [f'panel{i}']
    ws.cell(row=row+i,column=column+5).value = RF_LC3 [f'panel{i}']

#___RF strength___


for LoadCase in maindir:
    row =17
    for LoadTable in maindir[LoadCase]:
        if 'vonMieses' in LoadTable :
            for Panel in maindir[LoadCase] [LoadTable]:
                for EementID,Elementkey in  maindir[LoadCase] [LoadTable] [Panel].items():
                    if LoadCase == 'LC1': modifier = 2
                    elif LoadCase == 'LC2' : modifier = 5
                    elif LoadCase == 'LC3' : modifier = 8
                    ws.cell(row=row,column=modifier).value =  Elementkey ['RF_strength']
                    row +=1

for LoadCase in maindir:
    row =47
    for LoadTable in maindir[LoadCase]:
        if 'Stringer' in LoadTable :
            for Panel in maindir[LoadCase] [LoadTable]:
                j = 0
                for EementID,Elementkey in  maindir[LoadCase] [LoadTable] [Panel].items():
                    if j < 3:
                        if LoadCase == 'LC1': modifier = 2
                        elif LoadCase == 'LC2' : modifier = 5
                        elif LoadCase == 'LC3' : modifier = 8
                        ws.cell(row=row,column=modifier).value =  Elementkey ['RF_strength']
                        row +=1
                        j += 1


wb.save(result_file)
print('hey')

